import openpyxl
import json
from datetime import datetime

def excel_to_json_template(excel_path, json_output_path):
    """Convert Excel workbook to a structured JSON template."""
    
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    template = {
        "metadata": {
            "source_file": excel_path,
            "generated_at": datetime.now().isoformat(),
            "sheets": {}
        }
    }
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_data = {
            "sheet_name": sheet_name,
            "dimensions": {
                "max_row": ws.max_row,
                "max_column": ws.max_column
            },
            "cells": {},
            "merged_cells": [str(merged_range) for merged_range in ws.merged_cells.ranges],
            "column_dimensions": {},
            "row_dimensions": {}
        }
        
        # Capture column dimensions (width, hidden)
        for col_letter, col_dim in ws.column_dimensions.items():
            sheet_data["column_dimensions"][col_letter] = {
                "width": col_dim.width,
                "hidden": col_dim.hidden
            }
        
        # Capture row dimensions (height, hidden)
        for row_num, row_dim in ws.row_dimensions.items():
            sheet_data["row_dimensions"][str(row_num)] = {
                "height": row_dim.height,
                "hidden": row_dim.hidden
            }
        
        # Extract cell data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cell_key = f"{cell.column_letter}{cell.row}"
                    cell_info = {
                        "value": cell.value,
                        "data_type": cell.data_type,
                        "style": {
                            "font": {
                                "name": cell.font.name,
                                "size": cell.font.size,
                                "bold": cell.font.bold,
                                "italic": cell.font.italic,
                                "color": str(cell.font.color) if cell.font.color else None
                            },
                            "fill": {
                                "fill_type": cell.fill.fill_type,
                                "start_color": str(cell.fill.start_color) if cell.fill.start_color else None,
                                "end_color": str(cell.fill.end_color) if cell.fill.end_color else None
                            },
                            "border": {
                                "left": str(cell.border.left.style) if cell.border.left else None,
                                "right": str(cell.border.right.style) if cell.border.right else None,
                                "top": str(cell.border.top.style) if cell.border.top else None,
                                "bottom": str(cell.border.bottom.style) if cell.border.bottom else None
                            },
                            "alignment": {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                                "wrap_text": cell.alignment.wrap_text
                            },
                            "number_format": cell.number_format,
                            "protection": {
                                "locked": cell.protection.locked,
                                "hidden": cell.protection.hidden
                            }
                        }
                    }
                    sheet_data["cells"][cell_key] = cell_info
        
        template["metadata"]["sheets"][sheet_name] = sheet_data
    
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(template, f, indent=2, default=str)
    
    print(f"JSON template generated at: {json_output_path}")
    return template

def json_to_excel_template(json_path, excel_output_path):
    """Convert a structured JSON template to an Excel workbook."""

    with open(json_path, 'r', encoding='utf-8') as f:
        template = json.load(f)
    
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove default sheet
    
    for sheet_name, sheet_data in template["metadata"]["sheets"].items():
        ws = workbook.create_sheet(title=sheet_name)
        
        # Set column dimensions
        for col_letter, col_dim in sheet_data.get("column_dimensions", {}).items():
            ws.column_dimensions[col_letter].width = col_dim.get("width", 8.43)
            ws.column_dimensions[col_letter].hidden = col_dim.get("hidden", False)
        
        # Set row dimensions
        for row_num_str, row_dim in sheet_data.get("row_dimensions", {}).items():
            row_num = int(row_num_str)
            ws.row_dimensions[row_num].height = row_dim.get("height", 15)
            ws.row_dimensions[row_num].hidden = row_dim.get("hidden", False)
        
        # Set cell data and styles
        for cell_key, cell_info in sheet_data.get("cells", {}).items():
            cell = ws[cell_key]
            cell.value = cell_info.get("value")
            cell.data_type = cell_info.get("data_type")
            
            style_info = cell_info.get("style", {})
            font_info = style_info.get("font", {})
            fill_info = style_info.get("fill", {})
            border_info = style_info.get("border", {})
            alignment_info = style_info.get("alignment", {})
            
            cell.font = openpyxl.styles.Font(
                name=font_info.get("name"),
                size=font_info.get("size"),
                bold=font_info.get("bold"),
                italic=font_info.get("italic"),
                color=font_info.get("color")
            )
            
            cell.fill = openpyxl.styles.PatternFill(
                fill_type=fill_info.get("fill_type"),
                start_color=fill_info.get("start_color"),
                end_color=fill_info.get("end_color")
            )
            
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_info.get("left")),
                right=openpyxl.styles.Side(style=border_info.get("right")),
                top=openpyxl.styles.Side(style=border_info.get("top")),
                bottom=openpyxl.styles.Side(style=border_info.get("bottom"))
            )
            
            cell

if __name__ == "__main__":
    excel_path = r"C:\Users\ortho\.openclaw\workspace\TempFiles\Imported_Files\excel_input_name.xlsx"
    json_output_path = r"C:\Users\ortho\.openclaw\workspace\TempFiles\Exported_Files\excel_input_name.json"
    
    template = excel_to_json_template(excel_path, json_output_path)
    print(f"Processed {len(template['metadata']['sheets'])} sheets.")
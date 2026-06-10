import openpyxl
import json
from datetime import datetime

def excel_to_json_template(excel_path, json_output_path):
    """Convert Excel workbook to a structured JSON template."""
    
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    template = {
        "metadata": {
            "source_file": excel_path,
            "generated_at": datetime.now().isoformat(),
            "sheets": {}
        }
    }
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_data = {
            "sheet_name": sheet_name,
            "dimensions": {
                "max_row": ws.max_row,
                "max_column": ws.max_column
            },
            "cells": {},
            "merged_cells": [str(merged_range) for merged_range in ws.merged_cells.ranges],
            "column_dimensions": {},
            "row_dimensions": {}
        }
        
        # Capture column dimensions (width, hidden)
        for col_letter, col_dim in ws.column_dimensions.items():
            sheet_data["column_dimensions"][col_letter] = {
                "width": col_dim.width,
                "hidden": col_dim.hidden
            }
        
        # Capture row dimensions (height, hidden)
        for row_num, row_dim in ws.row_dimensions.items():
            sheet_data["row_dimensions"][str(row_num)] = {
                "height": row_dim.height,
                "hidden": row_dim.hidden
            }
        
        # Extract cell data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cell_key = f"{cell.column_letter}{cell.row}"
                    cell_info = {
                        "value": cell.value,
                        "data_type": cell.data_type,
                        "style": {
                            "font": {
                                "name": cell.font.name,
                                "size": cell.font.size,
                                "bold": cell.font.bold,
                                "italic": cell.font.italic,
                                "color": str(cell.font.color) if cell.font.color else None
                            },
                            "fill": {
                                "fill_type": cell.fill.fill_type,
                                "start_color": str(cell.fill.start_color) if cell.fill.start_color else None,
                                "end_color": str(cell.fill.end_color) if cell.fill.end_color else None
                            },
                            "border": {
                                "left": str(cell.border.left.style) if cell.border.left else None,
                                "right": str(cell.border.right.style) if cell.border.right else None,
                                "top": str(cell.border.top.style) if cell.border.top else None,
                                "bottom": str(cell.border.bottom.style) if cell.border.bottom else None
                            },
                            "alignment": {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                                "wrap_text": cell.alignment.wrap_text
                            },
                            "number_format": cell.number_format,
                            "protection": {
                                "locked": cell.protection.locked,
                                "hidden": cell.protection.hidden
                            }
                        }
                    }
                    sheet_data["cells"][cell_key] = cell_info
        
        template["metadata"]["sheets"][sheet_name] = sheet_data
    
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(template, f, indent=2, default=str)
    
    print(f"JSON template generated at: {json_output_path}")
    return template

if __name__ == "__main__":
    excel_path = r"C:\Users\ortho\.openclaw\workspace\PortalX LITE\Settings\PortalX_LITE_v5.xlsx"
    json_output_path = r"C:\Users\ortho\.openclaw\workspace\PortalX LITE\Settings\PortalX_LITE_v5_template.json"
    
    template = excel_to_json_template(excel_path, json_output_path)
    print(f"Processed {len(template['metadata']['sheets'])} sheets.")
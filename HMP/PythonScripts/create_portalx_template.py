import openpyxl
import json
from datetime import datetime

def create_portalx_template(excel_path, json_output_path):
    """Create a clean PortalX LITE template without sample data."""
    
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    template = {
        "metadata": {
            "source_file": excel_path,
            "generated_at": datetime.now().isoformat(),
            "description": "PortalX LITE template - headers and formatting only, no sample data",
            "sheets": {}
        }
    }
    
    for sheet_name in workbook.sheetnames:
        # Skip sheets with "sample" in the name
        if "sample" in sheet_name.lower():
            continue
            
        ws = workbook[sheet_name]
        sheet_template = {
            "sheet_name": sheet_name,
            "dimensions": {
                "max_row": ws.max_row,
                "max_column": ws.max_column
            },
            "header_rows": {},      # Header rows with values & formatting
            "data_start_row": None, # Will be set based on where data begins
            "formatting_rules": {}, # Rules for applying colors to sub-orders, etc.
            "column_dimensions": {},
            "row_dimensions": {},
            "merged_cells": [str(merged_range) for merged_range in ws.merged_cells.ranges]
        }
        
        # Capture column/row dimensions
        for col_letter, col_dim in ws.column_dimensions.items():
            sheet_template["column_dimensions"][col_letter] = {
                "width": col_dim.width,
                "hidden": col_dim.hidden
            }
        
        for row_num, row_dim in ws.row_dimensions.items():
            sheet_template["row_dimensions"][str(row_num)] = {
                "height": row_dim.height,
                "hidden": row_dim.hidden
            }
        
        # Identify header rows: assume row 1 is header, and any row with empty cells after column 1
        header_rows = []
        data_rows = []
        for row in ws.iter_rows():
            row_num = row[0].row
            # Check if row is empty (all cells None) - skip
            if all(cell.value is None for cell in row):
                continue
            # If row has text in column A, it's likely a data row
            if row[0].value is not None and row_num > 1:
                data_rows.append(row_num)
            else:
                header_rows.append(row_num)
        
        sheet_template["data_start_row"] = min(data_rows) if data_rows else None
        
        # Extract header rows only
        for row_num in header_rows:
            row_cells = {}
            for cell in ws[row_num]:
                cell_key = f"{cell.column_letter}{cell.row}"
                cell_info = {
                    "value": cell.value,
                    "data_type": cell.data_type,
                    "style": {
                        "font": {
                            "name": cell.font.name,
                            "size": cell.font.size,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "color": str(cell.font.color.rgb) if hasattr(cell.font.color, 'rgb') else str(cell.font.color)
                        },
                        "fill": {
                            "fill_type": cell.fill.fill_type,
                            "start_color": str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color),
                            "end_color": str(cell.fill.end_color.rgb) if hasattr(cell.fill.end_color, 'rgb') else str(cell.fill.end_color)
                        },
                        "border": {
                            "left": str(cell.border.left.style) if cell.border.left else None,
                            "right": str(cell.border.right.style) if cell.border.right else None,
                            "top": str(cell.border.top.style) if cell.border.top else None,
                            "bottom": str(cell.border.bottom.style) if cell.border.bottom else None
                        },
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        },
                        "number_format": cell.number_format,
                        "protection": {
                            "locked": cell.protection.locked,
                            "hidden": cell.protection.hidden
                        }
                    }
                }
                row_cells[cell_key] = cell_info
            sheet_template["header_rows"][str(row_num)] = row_cells
        
        # Detect formatting rules from sample data rows (if any exist before filtering)
        # This will capture things like "sub-orders have yellow background"
        formatting_rules = {}
        # You can manually define rules here based on your design:
        # Example: if a row has a specific fill color, note that as a rule
        sheet_template["formatting_rules"] = formatting_rules
        
        template["metadata"]["sheets"][sheet_name] = sheet_template
    
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(template, f, indent=2, default=str)
    
    print(f"Clean template generated at: {json_output_path}")
    print(f"Excluded sample sheets. Included {len(template['metadata']['sheets'])} sheets.")
    return template

if __name__ == "__main__":
    excel_path = r"C:\Users\ortho\.openclaw\workspace\PortalX LITE\Settings\PortalX_LITE_v5.xlsx"
    json_output_path = r"C:\Users\ortho\.openclaw\workspace\PortalX LITE\Settings\PortalX_LITE_template_clean.json"
    
    template = create_portalx_template(excel_path, json_output_path)